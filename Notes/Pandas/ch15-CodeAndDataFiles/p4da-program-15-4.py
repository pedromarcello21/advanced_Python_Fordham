# access and display all cells from a workbook
from openpyxl import load_workbook

# Read an .xlsx file
filename = 'PythonExcel.xlsx'
workbook = load_workbook(filename, data_only=True)

# access the list of sheets
sheetnames = workbook.sheetnames

for sheetname in sheetnames:
    print('=== Worksheet:', sheetname, '===')
    worksheet = workbook[sheetname]
    for row in worksheet.rows:
       for cell in row:
          if cell.value != None:
             print(cell.coordinate, cell.row, cell.column,
                   cell.data_type, cell.value, sep='|')
