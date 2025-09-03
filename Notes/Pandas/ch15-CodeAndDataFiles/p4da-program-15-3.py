# access and display cells of different types from a workbook
from openpyxl import load_workbook

# Read the .xlsx file
workbook = load_workbook('PythonExcel.xlsx', data_only=True)

# access the 'Sheet1' worksheet tab
worksheet = workbook['Sheet1']

# for cells in the worksheet, display their values and types
for row in range(1, 10):
    cell1 = worksheet.cell(row=row, column=1)
    cell2 = worksheet.cell(row=row, column=2)
    print(cell1.value, cell2.value, cell2.data_type,type(cell2.value), sep='|')
