# export spreadsheet data by locating row labels and column headings
from openpyxl import load_workbook

# for an .xlsx income statement, locate row and column labels
def getRowColCoords(worksheet, rowcodes, colcodes):
   rowdict = {}
   coldict = {}
   for row in worksheet.rows:
      for cell in row:
        if cell.value in rowcodes:
            rowdict[cell.value] = (cell.column, cell.row)
        elif cell.value in colcodes:
            coldict[cell.value] = (cell.column, cell.row)
   return rowdict, coldict

# given the row/col name locations, locate data and save to a file
def saveData(divname, worksheet, rowcodes, colcodes, rowdict, coldict, f):
   for rowname in rowcodes:
      rowix = rowdict[rowname][1]
      print(divname,rowname, sep='|', end='|', file=f)
      for colname in colcodes:
         colix = coldict[colname][0]
         cell = worksheet.cell(row=rowix, column=colix)
         print(cell.value, end='|', file=f)
      print(file=f)
   f.close()

divname = 'diva'
filename = 'incstmt-' + divname + '.xlsx'
workbook = load_workbook(filename, data_only=True)
worksheet = workbook['IncomeStatement']

rowcodes = ('SALE', 'CGS', 'SGA', 'ADV', 'DEP', 'RENT', 'OTHX')
colcodes = ('Act2019', 'Act2020', 'Proj2021')
rowdict, coldict = getRowColCoords(worksheet, rowcodes, colcodes)

f = open('incstmt- ' + divname + '.txt', 'w')
saveData(divname, worksheet, rowcodes, colcodes, rowdict, coldict, f)
